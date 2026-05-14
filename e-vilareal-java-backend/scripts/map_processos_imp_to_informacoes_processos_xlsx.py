#!/usr/bin/env python3
"""
Converte «Processos_imp.xls» (export do sistema legado) para o layout esperado por
InformacoesProcessosImportService: 1ª linha = cabeçalho; dados A–O a partir da
linha 2; colunas A=cliente (8 dígitos), B–F=autores, G–K=réus, L=proc., M=fase,
N=CNJ, O=descrição da ação.

Usa a aba «Relatório - Andamento Processos» (índice 1 por defeito), onde:
  - col 4 = N Pessoa Cliente
  - cols 6–10 = N Pessoa 1–5 Autor
  - cols 11–12, 15–17 = até 5 réus (N Pessoa 1–2 Réu e 3–5 Réu; o 6º réu col 18
    não cabe no layout legado de 5 colunas G–K)

Fases desconhecidas ou «INATIVO» na coluna Fase → coluna M vazia (import aceita).

Correcções de ID nas partes (autores/réus):
  - **9895 → 1510** — no legado «9895» era inexistente; na BD o Welton ficou em **1510**.

  Não se aplica **1510 → 1509**: isso era válido apenas quando 1510 ainda duplicava a NET; agora **1509** é só NET e **1510** é só Welton.

Não altera a coluna A (cliente do processo).

Dependências: pip install xlrd openpyxl
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from pathlib import Path


# (origem, destino) — aplicados em ordem; repetir até não haver mais substituições.
_REMAPEAR_PESSOA_PARTE_EM_ORDEM: tuple[tuple[int, int], ...] = ((9895, 1510),)


def _remap_parte_id(s: str) -> str:
    if not s or not s.isdigit():
        return s
    try:
        n = int(s)
    except ValueError:
        return s
    changed = True
    while changed:
        changed = False
        for a, b in _REMAPEAR_PESSOA_PARTE_EM_ORDEM:
            if n == a:
                n = b
                changed = True
    return str(n)


def _norm_id_texto(s: str) -> str:
    """«6272.0» ou «3918,0» → «6272» (ids de pessoa na planilha legada)."""
    t = (s or "").strip()
    if not t:
        return ""
    if re.fullmatch(r"-?\d+([.,]\d+)?", t):
        try:
            x = float(t.replace(",", "."))
            if x == int(x):
                return str(int(x))
        except ValueError:
            pass
    return t

# Alinhado a FasePlanilhaNormalizer (valores canónicos)
_FASE_CANON = {
    "em andamento": "Em Andamento",
    "aguardando documentos": "Ag. Documentos",
    "ag peticionar": "Ag. Peticionar",
    "aguardando peticionar": "Ag. Peticionar",
    "aguardando peticionamento": "Ag. Peticionar",
    "ag verificacao": "Ag. Verificação",
    "ag verificação": "Ag. Verificação",
    "aguardando verificação": "Ag. Verificação",
    "aguardando providencia": "Aguardando Providência",
    "aguardando providência": "Aguardando Providência",
    "protocolo": "Protocolo / Movimentação",
    "protocolo / movimentacao": "Protocolo / Movimentação",
    "movimentação": "Protocolo / Movimentação",
    "procedimento adm": "Procedimento Adm.",
    "procedimento administrativo": "Procedimento Adm.",
}


def _norm_fase_chave(s: str) -> str:
    t = s.strip().lower()
    t = unicodedata.normalize("NFD", t)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return " ".join(t.split())


def _fase_para_coluna_m(bruto: str) -> str:
    if not bruto or not str(bruto).strip():
        return ""
    ch = _norm_fase_chave(str(bruto))
    if ch in ("inativo",):
        return ""
    return _FASE_CANON.get(ch, "")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("entrada", type=Path, help="Caminho do Processos_imp.xls")
    ap.add_argument("saida", type=Path, help="Caminho do .xlsx de saída")
    ap.add_argument(
        "--sheet-index",
        type=int,
        default=1,
        help="Índice 0-based da aba fonte (default 1 = Relatório - Andamento Processos)",
    )
    args = ap.parse_args()

    try:
        import xlrd
        from openpyxl import Workbook
    except ImportError:
        print("Instale: pip install xlrd openpyxl", file=sys.stderr)
        return 1

    if not args.entrada.is_file():
        print(f"Ficheiro não encontrado: {args.entrada}", file=sys.stderr)
        return 1

    book = xlrd.open_workbook(str(args.entrada))
    if args.sheet_index < 0 or args.sheet_index >= book.nsheets:
        print("sheet-index inválido", file=sys.stderr)
        return 1
    sh = book.sheet_by_index(args.sheet_index)

    def cell_str(r: int, c: int) -> str:
        if c >= sh.ncols:
            return ""
        t = sh.cell_type(r, c)
        v = sh.cell_value(r, c)
        if t == xlrd.XL_CELL_EMPTY:
            return ""
        if t == xlrd.XL_CELL_NUMBER:
            if v == int(v):
                return str(int(v))
            return str(v).strip()
        if t == xlrd.XL_CELL_BOOLEAN:
            return "SIM" if v else "NÃO"
        return _norm_id_texto(str(v).strip() if v is not None else "")

    def codigo_oito(r: int, c: int) -> str:
        s = _norm_id_texto(cell_str(r, c))
        if not s:
            return ""
        if s.isdigit():
            return f"{int(s):08d}"
        return s

    wb = Workbook()
    out = wb.active
    out.title = "InformacoesProcessos"
    headers = [
        "Cliente (A)",
        "Autor1",
        "Autor2",
        "Autor3",
        "Autor4",
        "Autor5",
        "Reu1",
        "Reu2",
        "Reu3",
        "Reu4",
        "Reu5",
        "Proc (L)",
        "Fase (M)",
        "CNJ (N)",
        "Descricao (O)",
    ]
    for j, h in enumerate(headers):
        out.cell(row=1, column=j + 1, value=h)

    out_row = 2
    skipped = 0
    for r in range(1, sh.nrows):
        cli = cell_str(r, 4)
        proc = cell_str(r, 13)
        if not cli or not proc:
            skipped += 1
            continue
        cod_a = codigo_oito(r, 4)
        if not cod_a:
            skipped += 1
            continue

        autores = [_remap_parte_id(cell_str(r, c)) for c in (6, 7, 8, 9, 10)]
        reus_src_cols = (11, 12, 15, 16, 17)
        reus = [_remap_parte_id(cell_str(r, c)) for c in reus_src_cols]

        fase_m = _fase_para_coluna_m(cell_str(r, 14))
        cnj = cell_str(r, 19)
        desc = cell_str(r, 20)

        row_vals = [cod_a] + autores + reus + [proc, fase_m, cnj, desc]
        for j, val in enumerate(row_vals):
            out.cell(row=out_row, column=j + 1, value=val)
        out_row += 1

    args.saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(args.saida))
    print(
        f"OK: {args.saida} — linhas de dados: {out_row - 2} (ignoradas sem cliente/proc: {skipped}) "
        f"— folha fonte «{sh.name}»"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
