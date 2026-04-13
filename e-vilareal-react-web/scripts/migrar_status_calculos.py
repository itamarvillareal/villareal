#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Importação de status de cálculos aceitos a partir de Excel — gera SQL (INSERT … ON DUPLICATE KEY UPDATE)
que preserva titulos[] e parcelas[] já gravados ao atualizar chaves existentes (JSON_MERGE_PATCH).

Dependência: pip install openpyxl

Uso:
  python3 scripts/migrar_status_calculos.py --write-example ~/Downloads/exemplo_status_calculos.xlsx
  python3 scripts/migrar_status_calculos.py ~/Downloads/exemplo_status_calculos.xlsx
  python3 scripts/migrar_status_calculos.py ~/planilha.xlsx -o ~/Downloads/import_status_calculos.sql
  python3 scripts/migrar_status_calculos.py ~/planilha.xlsx --print   # imprime SQL no stdout (não grava)

Não executa nada no MySQL — apenas gera o ficheiro .sql (ou imprime) para revisão manual.
"""

from __future__ import annotations

import argparse
import base64
import json
import re
import sys
from copy import deepcopy
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuração — ajuste caminhos e nomes das colunas (cabeçalho na 1.ª linha)
# ---------------------------------------------------------------------------
EXCEL_PATH = Path(r"[CAMINHO DA PLANILHA]")
SHEET_NAME: str | None = None  # None = primeira aba
OUTPUT_SQL = Path.home() / "Downloads" / "import_status_calculos.sql"

COL_CLIENTE = "Código cliente"
COL_PROCESSO = "Processo"
COL_DIMENSAO = "Dimensão"
COL_ACEITO = "Aceito"
COL_DATA = "Data do cálculo aceito"

PANEL_CONFIG_DEFAULT = {
    "juros": "1 %",
    "multa": "10 %",
    "honorariosTipo": "fixos",
    "honorariosValor": "20 %",
    "honorariosVariaveisTexto": "",
    "indice": "INPC",
    "periodicidade": "mensal",
    "modeloListaDebitos": "01",
}

# Alinhado ao front (grade de títulos / parcelas por página)
N_TITULOS_VAZIOS_DEPOIS = 19
N_PARCELAS_VAZIAS_DEPOIS = 19
DEFAULT_TAXA_JUROS_PARCELAMENTO = "0,00"


def _norm_header(s: str) -> str:
    s = str(s or "").strip().lower()
    s = (
        s.replace("á", "a")
        .replace("ã", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ô", "o")
        .replace("ú", "u")
        .replace("ç", "c")
    )
    return re.sub(r"\s+", " ", s)


def _find_col(headers: list[str], name: str) -> int:
    """Localiza coluna pelo nome configurado (comparação normalizada)."""
    want = _norm_header(name)
    for i, h in enumerate(headers):
        if _norm_header(h) == want:
            return i
    return -1


def _pad_cliente8(val) -> str:
    n = int(re.sub(r"\D", "", str(val or "0")) or "0")
    if n < 1:
        n = 1
    return str(n).zfill(8)[:8]


def _int_processo(val) -> int:
    n = int(float(str(val).replace(",", ".")))
    return max(1, n)


def _int_dimensao(val) -> int:
    if val is None or str(val).strip() == "":
        return 0
    return max(0, int(float(str(val).replace(",", "."))))


def _linha_titulo_vazia() -> dict:
    return {
        "dataVencimento": "",
        "valorInicial": "",
        "atualizacaoMonetaria": "",
        "diasAtraso": "",
        "juros": "",
        "multa": "",
        "honorarios": "",
        "total": "",
        "descricaoValor": "",
        "datasEspeciais": None,
    }


def _linha_parcela_vazia() -> dict:
    return {
        "dataVencimento": "",
        "valorParcela": "",
        "honorariosParcela": "",
        "observacao": "",
        "dataPagamento": "",
    }


def _panel_config() -> dict:
    return deepcopy(PANEL_CONFIG_DEFAULT)


def _aceito_sim(val) -> bool:
    t = str(val or "").strip().upper()
    return t == "SIM"


def _normalizar_data_excel(val) -> str:
    if val is None or str(val).strip() == "":
        return ""
    if hasattr(val, "strftime"):
        return f"{val.day:02d}/{val.month:02d}/{val.year}"
    return str(val).strip()


def montar_payload(
    *,
    aceito: bool,
    data_calculo_br: str,
) -> dict:
    """
    aceito=True: parcelamentoAceito true; data em titulos[0].dataVencimento e parcelas[0].dataVencimento.
    aceito=False: parcelamentoAceito false; estrutura mínima sem valores monetários/datas preenchidos.
    """
    panel = _panel_config()
    tit0 = _linha_titulo_vazia()
    p0 = _linha_parcela_vazia()
    if aceito:
        d = str(data_calculo_br).strip()
        tit0["dataVencimento"] = d
        p0["dataVencimento"] = d

    titulos = [tit0] + [_linha_titulo_vazia() for _ in range(N_TITULOS_VAZIOS_DEPOIS)]
    parcelas = [p0] + [_linha_parcela_vazia() for _ in range(N_PARCELAS_VAZIAS_DEPOIS)]

    return {
        "pagina": 1,
        "paginaParcelamento": 1,
        "titulos": titulos,
        "parcelas": parcelas,
        "quantidadeParcelasInformada": "01",
        "taxaJurosParcelamento": DEFAULT_TAXA_JUROS_PARCELAMENTO,
        "limpezaAtiva": False,
        "snapshotAntesLimpeza": None,
        "cabecalho": {"autor": "", "reu": ""},
        "honorariosDataRecebimento": {},
        "parcelamentoAceito": aceito,
        "panelConfig": panel,
    }


def patch_para_duplicate(aceito: bool) -> dict:
    """Só chaves atualizadas em conflito — JSON_MERGE_PATCH não toca em titulos/parcelas."""
    return {
        "parcelamentoAceito": aceito,
        "panelConfig": _panel_config(),
    }


def _payload_b64_sql(payload: dict) -> str:
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    b64 = base64.standard_b64encode(raw.encode("utf-8")).decode("ascii")
    return f"CAST(CONVERT(FROM_BASE64('{b64}') USING utf8mb4) AS JSON)"


def sql_upsert_um(cod8: str, proc: int, dim: int, payload: dict, aceito: bool) -> str:
    """
    INSERT com payload completo; em duplicado faz merge só de parcelamentoAceito + panelConfig,
    preservando o restante do JSON já na linha (titulos, parcelas, etc.).
    Usa calculo_rodada.payload_json no primeiro argumento de JSON_MERGE_PATCH (valor existente).
    """
    ins_json = _payload_b64_sql(payload)
    patch = patch_para_duplicate(aceito)
    patch_b64 = _payload_b64_sql(patch)
    return f"""INSERT INTO calculo_rodada (codigo_cliente, numero_processo, dimensao, payload_json) VALUES
  ('{cod8}', {proc}, {dim}, {ins_json}) AS ins
ON DUPLICATE KEY UPDATE
  payload_json = JSON_MERGE_PATCH(
    calculo_rodada.payload_json,
    {patch_b64}
  );"""


def ler_planilha(path: Path, sheet_name: str | None) -> list[dict]:
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit("Instale openpyxl: pip install openpyxl") from e

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    headers = [str(c) if c is not None else "" for c in rows[0]]
    idx = {
        "codigo_cliente": _find_col(headers, COL_CLIENTE),
        "numero_processo": _find_col(headers, COL_PROCESSO),
        "dimensao": _find_col(headers, COL_DIMENSAO),
        "aceito": _find_col(headers, COL_ACEITO),
        "data": _find_col(headers, COL_DATA),
    }
    missing = [k for k, v in idx.items() if v < 0]
    if missing:
        raise SystemExit(
            "Cabeçalhos não encontrados para: "
            + ", ".join(missing)
            + "\nNomes configurados: "
            + f"{COL_CLIENTE!r}, {COL_PROCESSO!r}, {COL_DIMENSAO!r}, {COL_ACEITO!r}, {COL_DATA!r}"
            + "\nCabeçalhos lidos: "
            + repr(headers)
        )

    out: list[dict] = []
    for row in rows[1:]:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        out.append(
            {
                "codigo_cliente": row[idx["codigo_cliente"]],
                "numero_processo": row[idx["numero_processo"]],
                "dimensao": row[idx["dimensao"]] if idx["dimensao"] >= 0 else "",
                "aceito": row[idx["aceito"]],
                "data": row[idx["data"]] if idx["data"] >= 0 else "",
            }
        )
    return out


def processar_linhas(rows: list[dict]) -> list[tuple[str, int, int, dict, bool]]:
    """Lista de (cod8, proc, dim, payload, aceito) validada."""
    out: list[tuple[str, int, int, dict, bool]] = []
    for i, r in enumerate(rows, start=2):
        aceito = _aceito_sim(r["aceito"])
        data_br = _normalizar_data_excel(r["data"])
        if aceito and not data_br:
            raise SystemExit(f"Linha Excel {i}: «Aceito»=SIM exige «{COL_DATA}» preenchida.")
        if not aceito and data_br:
            # não é erro — ignoramos a data se não aceito
            data_br = ""
        cod8 = _pad_cliente8(r["codigo_cliente"])
        proc = _int_processo(r["numero_processo"])
        dim = _int_dimensao(r.get("dimensao", ""))
        payload = montar_payload(aceito=aceito, data_calculo_br=data_br)
        out.append((cod8, proc, dim, payload, aceito))
    return out


def gerar_sql(rows: list[dict]) -> str:
    blocos = [
        "SET NAMES utf8mb4;",
        "-- Importação de status (aceito / não aceito) — UPSERT por (codigo_cliente, numero_processo, dimensao).",
        "-- Em duplicado: JSON_MERGE_PATCH só aplica parcelamentoAceito + panelConfig; titulos/parcelas preservados.",
        "",
    ]
    for cod8, proc, dim, payload, aceito in processar_linhas(rows):
        blocos.append(f"-- cliente={cod8} proc={proc} dim={dim} aceito={aceito}")
        blocos.append(sql_upsert_um(cod8, proc, dim, payload, aceito))
        blocos.append("")
    return "\n".join(blocos).rstrip() + "\n"


def escrever_exemplo_excel(path: Path) -> None:
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit("Instale openpyxl: pip install openpyxl") from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "status"
    ws.append(
        [
            COL_CLIENTE,
            COL_PROCESSO,
            COL_DIMENSAO,
            COL_ACEITO,
            COL_DATA,
        ]
    )
    # 4 linhas de teste (conforme especificação)
    ws.append([10, 2, 0, "SIM", "06/10/2022"])
    ws.append([728, 199, 0, "SIM", "15/03/2021"])
    ws.append([922, 7, 0, "", ""])
    ws.append([29, 6, 0, "SIM", "01/06/2020"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Excel de exemplo gravado em: {path}")


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Gera SQL para importar status de cálculos aceitos (preserva titulos/parcelas em UPDATE)."
    )
    ap.add_argument("excel", nargs="?", help="Caminho do .xlsx (se omitido, usa EXCEL_PATH se existir)")
    ap.add_argument("-o", "--output", default=str(OUTPUT_SQL), help="Ficheiro SQL de saída")
    ap.add_argument(
        "--sheet",
        default=SHEET_NAME,
        help="Nome da aba (omitir = primeira aba, ou o valor de SHEET_NAME no script)",
    )
    ap.add_argument(
        "--write-example",
        metavar="PATH",
        help="Gera Excel de exemplo com 4 linhas em PATH e termina",
    )
    ap.add_argument(
        "--print",
        action="store_true",
        help="Imprime o SQL no stdout (além de gravar em -o, salvo se usar só --print com -o -)",
    )
    ap.add_argument(
        "--stdout-only",
        action="store_true",
        help="Só imprime SQL no stdout; não grava ficheiro",
    )
    args = ap.parse_args()

    if args.write_example:
        escrever_exemplo_excel(Path(args.write_example).expanduser())
        return

    xlsx: Path | None = None
    if args.excel:
        xlsx = Path(args.excel).expanduser()
    else:
        cand = EXCEL_PATH.expanduser() if isinstance(EXCEL_PATH, Path) else Path(EXCEL_PATH)
        if cand.is_file():
            xlsx = cand

    if not xlsx or not xlsx.is_file():
        ap.print_help()
        print(
            "\nIndique o Excel: argumento posicional ou defina EXCEL_PATH no topo do script.",
            file=sys.stderr,
        )
        sys.exit(2)

    rows = ler_planilha(xlsx, args.sheet)
    if not rows:
        sys.exit("Nenhuma linha de dados na planilha.")

    sql = gerar_sql(rows)
    if args.print or args.stdout_only:
        print(sql, end="")
    if not args.stdout_only:
        out = Path(args.output).expanduser()
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(sql, encoding="utf-8")
        print(f"SQL gerado: {out} ({len(rows)} linha(s))", file=sys.stderr if args.print else sys.stdout)


if __name__ == "__main__":
    main()
