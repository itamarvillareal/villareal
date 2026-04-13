#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera SQL (INSERT ... ON DUPLICATE KEY UPDATE) para `vilareal.calculo_rodada`
a partir de uma planilha Excel com cálculos já aceitos (payload mínimo + parcelamentoAceito).

Dependência: pip install openpyxl

Uso:
  python3 scripts/migrar_calculos_aceitos.py /caminho/planilha.xlsx
  python3 scripts/migrar_calculos_aceitos.py --write-sample ~/Downloads/exemplo_calculos_aceitos.xlsx
  python3 scripts/migrar_calculos_aceitos.py ~/Downloads/exemplo_calculos_aceitos.xlsx -o ~/Downloads/import_calculos_aceitos.sql
"""

from __future__ import annotations

import argparse
import base64
import json
import re
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuração — ajuste nomes das colunas na primeira linha do Excel
# (valores são comparados em minúsculas, sem acento extra onde indicado)
# ---------------------------------------------------------------------------
OUTPUT_SQL_DEFAULT = str(Path.home() / "Downloads" / "import_calculos_aceitos.sql")

# Nomes esperados no cabeçalho (primeira linha). O script aceita o primeiro que existir.
COL_CODIGO_CLIENTE = ("código do cliente", "codigo do cliente", "codigo_cliente", "cliente")
COL_NUMERO_PROCESSO = ("número do processo", "numero do processo", "numero_processo", "processo")
COL_DIMENSAO = ("dimensão", "dimensao", "dim")
COL_QTD_PARCELAS = ("quantidade de parcelas", "quantidade parcelas", "parcelas", "qtd parcelas")
COL_VALOR_TOTAL = ("valor total do cálculo", "valor total do calculo", "valor total", "total")
COL_DATA_BASE = ("data base do cálculo", "data base do calculo", "data base", "data")
COL_JUROS = ("taxa de juros", "juros")
COL_MULTA = ("taxa de multa", "multa")
COL_HON_TIPO = ("tipo de honorários", "tipo de honorarios", "honorarios tipo", "hon. tipo")
COL_HON_VALOR = ("valor de honorários", "valor de honorarios", "honorarios valor", "hon. valor")

# Taxa de juros do parcelamento (Price), não confundir com juros de mora do painel — default fixo.
DEFAULT_TAXA_JUROS_PARCELAMENTO = "0,00"

# Linhas vazias modelo (alinhado ao front: ~20 parcelas por página, títulos com grade longa)
N_PARCELAS_VAZIAS_DEPOIS = 19
N_TITULOS_VAZIOS_DEPOIS = 19


def _norm_header(s: str) -> str:
    s = str(s or "").strip().lower()
    s = (
        s.replace("á", "a")
        .replace("ã", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ô", "o")
        .replace("ú", "u")
        .replace("ç", "c")
    )
    return re.sub(r"\s+", " ", s)


def _find_col(headers: list[str], candidates: tuple[str, ...]) -> int:
    norm_map = {_norm_header(h): i for i, h in enumerate(headers)}
    for c in candidates:
        key = _norm_header(c)
        if key in norm_map:
            return norm_map[key]
    return -1


def _pad_cliente8(val) -> str:
    n = int(re.sub(r"\D", "", str(val or "0")) or "0")
    if n < 1:
        n = 1
    return str(n).zfill(8)[:8]


def _int_processo(val) -> int:
    n = int(float(str(val).replace(",", ".")))
    return max(1, n)


def _int_dimensao(val) -> int:
    if val is None or str(val).strip() == "":
        return 0
    return max(0, int(float(str(val).replace(",", "."))))


def _int_parcelas(val) -> int:
    s = re.sub(r"\D", "", str(val or "0"))
    n = int(s) if s else 0
    return min(9999, max(0, n))


def _decimal_valor(val) -> Decimal:
    if val is None or val == "":
        return Decimal("0")
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip()
    s = re.sub(r"r\$\s?", "", s, flags=re.I)
    s = s.replace(".", "").replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s:
        return Decimal("0")
    return Decimal(s)


def _format_brl(d: Decimal) -> str:
    q = d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    inteiro, frac = str(q).split(".") if "." in str(q) else (str(q), "00")
    # milhar pt-BR
    sign = ""
    if inteiro.startswith("-"):
        sign = "-"
        inteiro = inteiro[1:]
    rev = inteiro[::-1]
    parts = [rev[i : i + 3] for i in range(0, len(rev), 3)]
    mil = ".".join(p[::-1] for p in reversed(parts))
    return f"R$ {sign}{mil},{frac.zfill(2)}"


def _norm_pct_juros_multa(s: str) -> str:
    t = str(s or "").strip()
    if not t:
        return "0 %"
    if "%" not in t:
        t = t + " %"
    return t


def _norm_hon_tipo(s: str) -> str:
    t = _norm_header(s).replace(" ", "")
    if "vari" in t:
        return "variaveis"
    return "fixos"


def _norm_hon_valor_pct(s: str) -> str:
    t = str(s or "").strip()
    if not t:
        return "0 %"
    if "%" not in t:
        t = t + " %"
    return t


def _linha_titulo_vazia() -> dict:
    return {
        "dataVencimento": "",
        "valorInicial": "",
        "atualizacaoMonetaria": "",
        "diasAtraso": "",
        "juros": "",
        "multa": "",
        "honorarios": "",
        "total": "",
        "descricaoValor": "",
        "datasEspeciais": None,
    }


def _linha_parcela_vazia() -> dict:
    return {
        "dataVencimento": "",
        "valorParcela": "",
        "honorariosParcela": "",
        "observacao": "",
        "dataPagamento": "",
    }


def montar_payload(
    *,
    data_base_br: str,
    valor_total: Decimal,
    qtd_parcelas: int,
    juros: str,
    multa: str,
    honorarios_tipo: str,
    honorarios_valor: str,
    taxa_juros_parcelamento: str,
) -> dict:
    brl = _format_brl(valor_total)
    qtd_str = str(qtd_parcelas) if qtd_parcelas > 99 else str(qtd_parcelas).zfill(2)

    tit0 = _linha_titulo_vazia()
    tit0["dataVencimento"] = str(data_base_br).strip()
    tit0["valorInicial"] = brl
    tit0["total"] = brl

    p0 = _linha_parcela_vazia()
    p0["dataVencimento"] = str(data_base_br).strip()
    p0["dataPagamento"] = str(data_base_br).strip()
    p0["valorParcela"] = brl
    p0["honorariosParcela"] = ""

    titulos = [tit0] + [_linha_titulo_vazia() for _ in range(N_TITULOS_VAZIOS_DEPOIS)]
    parcelas = [p0] + [_linha_parcela_vazia() for _ in range(N_PARCELAS_VAZIAS_DEPOIS)]

    return {
        "pagina": 1,
        "paginaParcelamento": 1,
        "titulos": titulos,
        "parcelas": parcelas,
        "quantidadeParcelasInformada": qtd_str,
        "taxaJurosParcelamento": taxa_juros_parcelamento,
        "limpezaAtiva": False,
        "snapshotAntesLimpeza": None,
        "cabecalho": {"autor": "", "reu": ""},
        "honorariosDataRecebimento": {},
        "parcelamentoAceito": True,
        "panelConfig": {
            "juros": _norm_pct_juros_multa(juros),
            "multa": _norm_pct_juros_multa(multa),
            "honorariosTipo": honorarios_tipo,
            "honorariosValor": _norm_hon_valor_pct(honorarios_valor),
            "honorariosVariaveisTexto": "",
            "indice": "INPC",
            "periodicidade": "mensal",
            "modeloListaDebitos": "01",
        },
    }


def sql_row_payload(cod8: str, proc: int, dim: int, payload: dict) -> str:
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    b64 = base64.standard_b64encode(raw.encode("utf-8")).decode("ascii")
    return (
        f"('{cod8}', {proc}, {dim}, CAST(CONVERT(FROM_BASE64('{b64}') USING utf8mb4) AS JSON))"
    )


def ler_planilha(path: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit("Instale openpyxl: pip install openpyxl") from e

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(c) if c is not None else "" for c in rows[0]]
    idx_cod = _find_col(headers, COL_CODIGO_CLIENTE)
    idx_proc = _find_col(headers, COL_NUMERO_PROCESSO)
    idx_dim = _find_col(headers, COL_DIMENSAO)
    idx_qtd = _find_col(headers, COL_QTD_PARCELAS)
    idx_tot = _find_col(headers, COL_VALOR_TOTAL)
    idx_dt = _find_col(headers, COL_DATA_BASE)
    idx_j = _find_col(headers, COL_JUROS)
    idx_m = _find_col(headers, COL_MULTA)
    idx_ht = _find_col(headers, COL_HON_TIPO)
    idx_hv = _find_col(headers, COL_HON_VALOR)
    needed = {
        "codigo_cliente": idx_cod,
        "numero_processo": idx_proc,
        "quantidade_parcelas": idx_qtd,
        "valor_total": idx_tot,
        "data_base": idx_dt,
        "juros": idx_j,
        "multa": idx_m,
        "hon_tipo": idx_ht,
        "hon_valor": idx_hv,
    }
    missing = [k for k, v in needed.items() if v < 0]
    if missing:
        raise SystemExit(
            "Cabeçalhos não encontrados para: "
            + ", ".join(missing)
            + "\nCabeçalhos lidos: "
            + repr(headers)
        )

    out = []
    for row in rows[1:]:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        out.append(
            {
                "codigo_cliente": row[idx_cod],
                "numero_processo": row[idx_proc],
                "dimensao": row[idx_dim] if idx_dim >= 0 else "",
                "quantidade_parcelas": row[idx_qtd],
                "valor_total": row[idx_tot],
                "data_base": row[idx_dt],
                "juros": row[idx_j],
                "multa": row[idx_m],
                "hon_tipo": row[idx_ht],
                "hon_valor": row[idx_hv],
            }
        )
    return out


def gerar_sql_corrigido(rows: list[dict], taxa_parcelamento: str) -> str:
    """Um único INSERT ... VALUES ... ON DUPLICATE KEY UPDATE (sintaxe MySQL)."""
    parts = [
        "SET NAMES utf8mb4;",
        "-- UPSERT pela UNIQUE (codigo_cliente, numero_processo, dimensao).",
        "",
        "INSERT INTO calculo_rodada (codigo_cliente, numero_processo, dimensao, payload_json) VALUES",
    ]
    vals = []
    for r in rows:
        cod8 = _pad_cliente8(r["codigo_cliente"])
        proc = _int_processo(r["numero_processo"])
        dim = _int_dimensao(r["dimensao"])
        qtd = _int_parcelas(r["quantidade_parcelas"])
        total = _decimal_valor(r["valor_total"])
        data_base = r["data_base"]
        if hasattr(data_base, "strftime"):
            data_base = f"{data_base.day:02d}/{data_base.month:02d}/{data_base.year}"
        else:
            data_base = str(data_base or "").strip()
        payload = montar_payload(
            data_base_br=data_base,
            valor_total=total,
            qtd_parcelas=qtd,
            juros=str(r["juros"] or ""),
            multa=str(r["multa"] or ""),
            honorarios_tipo=_norm_hon_tipo(str(r["hon_tipo"] or "fixos")),
            honorarios_valor=str(r["hon_valor"] or ""),
            taxa_juros_parcelamento=taxa_parcelamento,
        )
        vals.append(sql_row_payload(cod8, proc, dim, payload))
    parts.append(",\n".join(vals))
    parts.append("ON DUPLICATE KEY UPDATE payload_json = VALUES(payload_json);")
    return "\n".join(parts)


def escrever_exemplo_excel(path: Path) -> None:
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit("Instale openpyxl: pip install openpyxl") from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "aceitos"
    headers = [
        "Código do cliente",
        "Número do processo",
        "Dimensão",
        "Quantidade de parcelas",
        "Valor total do cálculo",
        "Data base do cálculo",
        "Taxa de juros",
        "Taxa de multa",
        "Tipo de honorários",
        "Valor de honorários",
    ]
    ws.append(headers)
    # 3 linhas fictícias (conforme pedido do utilizador)
    ws.append([10201, 15, 0, 12, 5000, "06/10/2022", "1 %", "10 %", "fixos", "20 %"])
    ws.append([72819, 90, 3, 6, 8500, "15/03/2021", "1 %", "10 %", "fixos", "20 %"])
    ws.append([92270, 12, 3, 24, 2000, "01/06/2020", "1 %", "10 %", "variaveis", "15 %"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Excel de exemplo gravado em: {path}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Gera SQL import_calculos_aceitos.sql a partir do Excel.")
    ap.add_argument("excel", nargs="?", help="Caminho do .xlsx")
    ap.add_argument("-o", "--output", default=OUTPUT_SQL_DEFAULT, help="Ficheiro SQL de saída")
    ap.add_argument(
        "--taxa-parcelamento",
        default=DEFAULT_TAXA_JUROS_PARCELAMENTO,
        help='Taxa Price mensal no campo taxaJurosParcelamento (default "0,00")',
    )
    ap.add_argument(
        "--write-sample",
        metavar="PATH",
        help="Gera um Excel de exemplo em PATH e termina",
    )
    args = ap.parse_args()

    if args.write_sample:
        escrever_exemplo_excel(Path(args.write_sample).expanduser())
        return

    if not args.excel:
        ap.print_help()
        sys.exit(2)

    xlsx = Path(args.excel).expanduser()
    if not xlsx.is_file():
        sys.exit(f"Ficheiro não encontrado: {xlsx}")

    rows = ler_planilha(xlsx)
    if not rows:
        sys.exit("Nenhuma linha de dados na planilha.")

    sql = gerar_sql_corrigido(rows, args.taxa_parcelamento)
    out = Path(args.output).expanduser()
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(sql, encoding="utf-8")
    print(f"SQL gerado: {out} ({len(rows)} linha(s))")


if __name__ == "__main__":
    main()
